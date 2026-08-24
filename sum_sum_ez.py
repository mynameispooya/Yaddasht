import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta
import io
import os

st.set_page_config(page_title="سوپر اپلیکیشن محاسبه و تجمیع اضافه کاری (گروهی)", layout="wide")

st.title("🚀 سوپر اپلیکیشن مدیریت و تجمیع اضافه کاری پرسنل")
st.write("با استفاده از این سیستم می‌توانید بخش‌های مختلف ایجاد کرده، برای هر بخش فایل تردد و فایل گروه کاری را آپلود کنید تا محاسبات دقیق بر اساس کدهای مشترک انجام شود.")

# ----------------------------------------------------
# توابع محاسباتی و پردازشی
# ----------------------------------------------------

def parse_time(time_str):
    """تبدیل رشته ساعت به شیء datetime با پشتیبانی از اعداد فارسی و انگلیسی"""
    if not time_str or pd.isna(time_str) or str(time_str).strip() in ['-', '']:
        return None
    
    clean_str = str(time_str).replace('\u200f', '').replace('\u200e', '').strip()
    
    persian_digits = '۰۱۲۳۴۵۶۷۸۹'
    english_digits = '0123456789'
    translation_table = str.maketrans(persian_digits, english_digits)
    clean_str = clean_str.translate(translation_table)
    
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(clean_str, fmt)
        except ValueError:
            pass
    return None

def calculate_duration_minutes(entry_str, exit_str):
    """محاسبه دقیق کارکرد به دقیقه با اعمال قانون شروع رسمی (07:15)"""
    entry_dt = parse_time(entry_str)
    exit_dt = parse_time(exit_str)
    
    if not entry_dt or not exit_dt:
        return 0
    
    # ساعت مبنای رسمی شرکت: 07:15
    official_start = entry_dt.replace(hour=7, minute=15, second=0)
    
    # قانون 07:15: اگر ورود زودتر باشد، مبنا 07:15 است
    effective_entry = max(entry_dt, official_start)
    
    if exit_dt < effective_entry:
        exit_dt += timedelta(days=1)
        
    duration = exit_dt - effective_entry
    return int(duration.total_seconds() // 60)

def minutes_to_hhmm(total_minutes):
    """تبدیل دقیقه به فرمت HH:MM"""
    if total_minutes <= 0:
        return "00:00"
    hours = total_minutes // 60
    minutes = total_minutes % 60
    return f"{hours:02d}:{minutes:02d}"

def extract_allowed_codes_from_group_excel(group_file_bytes):
    """استخراج مجموعه کدهای پرسنلی مجاز از فایل group.xlsx"""
    wb = openpyxl.load_workbook(io.BytesIO(group_file_bytes), data_only=True)
    sheet = wb.active
    
    allowed_codes = set()
    code_col = None
    header_row = None
    
    # پیدا کردن هوشمند ستون "کد"
    for r in range(1, min(30, sheet.max_row + 1)):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(row=r, column=c).value
            if val and str(val).strip() == 'کد':
                header_row = r
                code_col = c
                break
        if code_col:
            break
            
    if code_col and header_row:
        for r in range(header_row + 1, sheet.max_row + 1):
            val = sheet.cell(row=r, column=code_col).value
            if val is not None and str(val).strip() != '':
                code_clean = str(val).strip().replace('\u200f', '').replace('\u200e', '')
                allowed_codes.add(code_clean)
                
    return allowed_codes

def process_section_data(main_file_bytes, group_file_bytes):
    """پردازش فایل اصلی و فیلتر بر اساس کدهای موجود در فایل گروه"""
    allowed_codes = extract_allowed_codes_from_group_excel(group_file_bytes)
    
    wb_main = openpyxl.load_workbook(io.BytesIO(main_file_bytes), data_only=True)
    sheet = wb_main.active
    
    records = []
    for r in range(15, sheet.max_row + 1):
        p_code = sheet.cell(row=r, column=20).value
        name = sheet.cell(row=r, column=19).value
        entry = sheet.cell(row=r, column=10).value
        exit_time = sheet.cell(row=r, column=5).value
        
        if p_code or name:
            p_code_str = str(p_code).strip().replace('\u200f', '').replace('\u200e', '') if p_code is not None else '-'
            
            # **شرکت دادن فقط کدهای مشترک با فایل گروه**
            if p_code_str in allowed_codes:
                name_str = str(name).strip() if name is not None else '-'
                entry_str = str(entry).strip() if entry is not None else '-'
                exit_str = str(exit_time).strip() if exit_time is not None else '-'
                
                dur_min = calculate_duration_minutes(entry_str, exit_str)
                dur_hhmm = minutes_to_hhmm(dur_min)
                
                records.append({
                    'کد پرسنلی': p_code_str,
                    'نام و نام خانوادگی': name_str,
                    'ورود': entry_str,
                    'خروج': exit_str,
                    'کارکرد (دقیقه)': dur_min,
                    'مدت حضور (کارکرد)': dur_hhmm
                })
                
    df = pd.DataFrame(records)
    return df, len(allowed_codes)

def generate_styled_master_excel(master_df, section_names):
    """تولید فایل اکسل نهایی تجمیعی با هایلایت و استایل رسمی"""
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "گزارش تجمیعی کل"
    ws.views.sheetView[0].rightToLeft = True  # راست به چپ
    
    headers = list(master_df.columns)
    ws.append(headers)
    
    # استایل هدر
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # استایل‌های بدنه و هایلایت
    highlight_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # سبز ملایم
    regular_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", size=10, bold=True)
    
    for row_idx, row_data in enumerate(master_df.values, 2):
        has_work = False
        total_work_val = row_data[-1]  # ستون آخر "مجموع اضافه کار کلی"
        
        if total_work_val != "00:00" and total_work_val != "-":
            has_work = True
            
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = bold_font if col_idx in [1, 2, len(row_data)] else regular_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
            # هایلایت سطر در صورت داشتن کارکرد
            if has_work:
                cell.fill = highlight_fill
                
    # تنظیم عرض ستون‌ها
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 16)
        
    wb.save(output)
    return output.getvalue()

# ----------------------------------------------------
# مدیریت وضعیت نشست (Session State) برای بخش‌ها
# ----------------------------------------------------

if 'num_sections' not in st.session_state:
    st.session_state.num_sections = 1

def add_section():
    st.session_state.num_sections += 1

def remove_section():
    if st.session_state.num_sections > 1:
        st.session_state.num_sections -= 1

# ----------------------------------------------------
# رابط کاربری Streamlit
# ----------------------------------------------------

st.subheader("📁 بخش‌های آپلود فایل")

sections_data = {}
all_personnel_info = {}

for sec_idx in range(1, st.session_state.num_sections + 1):
    with st.expander(f"📌 بخش {sec_idx}", expanded=True):
        col1, col2 = st.columns(2)
        
        with col1:
            main_file = st.file_uploader(
                f"فایل اصلی تردد (file.xlsx) - بخش {sec_idx}",
                type=["xlsx", "xls"],
                key=f"main_{sec_idx}"
            )
            
        with col2:
            group_file = st.file_uploader(
                f"فایل اعضای گروه (group.xlsx) - بخش {sec_idx}",
                type=["xlsx", "xls"],
                key=f"group_{sec_idx}"
            )
            
        if main_file and group_file:
            try:
                df_sec, num_allowed = process_section_data(main_file.read(), group_file.read())
                sections_data[f"بخش {sec_idx}"] = df_sec
                
                # ثبت اطلاعات پرسنل
                for _, r in df_sec.iterrows():
                    p_code = r['کد پرسنلی']
                    if p_code != '-' and p_code not in all_personnel_info:
                        all_personnel_info[p_code] = r['نام و نام خانوادگی']
                        
                st.success(f"✅ پردازش بخش {sec_idx} موفقیت‌آمیز بود! (تعداد اعضای گروه: {num_allowed} | تعداد تطبیق‌یافته: {len(df_sec)})")
                
                # دانلود خروجی همین بخش
                output_sec = io.BytesIO()
                with pd.ExcelWriter(output_sec, engine='openpyxl') as writer:
                    df_sec.drop(columns=['کارکرد (دقیقه)']).to_excel(writer, index=False)
                    
                st.download_button(
                    label=f"📥 دانلود خروجی بخش {sec_idx}",
                    data=output_sec.getvalue(),
                    file_name=f"خروجی_محاسبات_بخش_{sec_idx}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_sec_{sec_idx}"
                )
            except Exception as e:
                st.error(f"خطا در پردازش بخش {sec_idx}: {e}")

# دکمه‌های افزودن / حذف بخش
col_add, col_rem, _ = st.columns([1, 1, 4])
with col_add:
    st.button("➕ افزودن بخش جدید", on_click=add_section, use_container_width=True)
with col_rem:
    if st.session_state.num_sections > 1:
        st.button("➖ حذف آخرین بخش", on_click=remove_section, use_container_width=True)

# ----------------------------------------------------
# ادغام نهایی (Merge) و نمایش گزارش کل
# ----------------------------------------------------

if sections_data:
    st.divider()
    st.subheader("📊 گزارش تجمیعی کل (Merge کامل بر اساس کدهای مشترک)")
    
    all_codes_sorted = sorted(list(all_personnel_info.keys()))
    section_names = list(sections_data.keys())
    
    master_rows = []
    
    for p_code in all_codes_sorted:
        row_dict = {
            'کد پرسنلی': p_code,
            'نام و نام خانوادگی': all_personnel_info.get(p_code, '-')
        }
        
        total_minutes = 0
        
        for sec_name in section_names:
            df_sec = sections_data[sec_name]
            match = df_sec[df_sec['کد پرسنلی'] == p_code]
            
            if not match.empty:
                dur_min = match.iloc[0]['کارکرد (دقیقه)']
                dur_str = match.iloc[0]['مدت حضور (کارکرد)']
                total_minutes += dur_min
                row_dict[f"اضافه کاری ({sec_name})"] = dur_str
            else:
                row_dict[f"اضافه کاری ({sec_name})"] = "-"
                
        row_dict['مجموع اضافه کار کلی'] = minutes_to_hhmm(total_minutes)
        master_rows.append(row_dict)
        
    master_df = pd.DataFrame(master_rows)
    
    # نمایش جدول تجمیعی
    st.dataframe(master_df, use_container_width=True)
    
    # دانلود فایل اکسل تجمیعی نهایی با استایل و هایلایت
    master_excel_bytes = generate_styled_master_excel(master_df, section_names)
    
    st.download_button(
        label="🔥 📥 دانلود فایل اکسل تجمیعی کل (با هایلایت و فرمت رسمی)",
        data=master_excel_bytes,
        file_name="گزارش_تجمیعی_اضافه_کاری_کل_بخش‌ها.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
