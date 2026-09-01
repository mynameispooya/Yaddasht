import streamlit as st
import pandas as pd
import io
from openpyxl.styles import PatternFill

st.set_page_config(page_title="پردازش تردد پرسنل", layout="wide")

st.title("اپلیکیشن پردازش و تفکیک فایل‌های تردد")
st.write("فایل اکسل خروجی دستگاه (حاوی ستون‌های multi-index) را آپلود کنید تا ردیف‌های دارای نقص تردد شناسایی و تفکیک شوند.")

uploaded_file = st.file_uploader("آپلود فایل اکسل", type=["xlsx"])

if uploaded_file is not None:
    try:
        # ۱. پاکسازی اولیه و رفع ساختار Multi-Index
        df_raw = pd.read_excel(uploaded_file, header=None)
        
        # یافتن سطر شروع داده‌ها به صورت داینامیک
        start_indices = df_raw.index[df_raw.isin(['اطلاعات تردد']).any(axis=1)].tolist()
        if not start_indices:
            st.error("ساختار فایل معتبر نیست (سلول 'اطلاعات تردد' یافت نشد).")
            st.stop()
            
        start_idx = start_indices[0] + 2
        df_clean = df_raw.iloc[start_idx:].copy()
        
        # استخراج ستون‌های ورود، خروج، گیت‌ها، تاریخ و مشخصات پرسنلی
        df_clean = df_clean[[3, 4, 6, 9, 17, 18, 19]].copy()
        df_clean.columns = ['گیت خروج', 'خروج', 'گیت ورود', 'ورود', 'تاریخ', 'نام و نام خانوادگی', 'کد پرسنلی']
        
        # حذف ردیف‌های اضافی "مجموع"
        df_clean = df_clean[~df_clean['تاریخ'].astype(str).str.contains('مجم')]
        
        # ۲. نظیر کردن کدهای پرسنلی و تاریخ‌ها به ردیف‌های خالی
        df_clean['نام و نام خانوادگی'] = df_clean['نام و نام خانوادگی'].ffill()
        df_clean['کد پرسنلی'] = df_clean['کد پرسنلی'].ffill()
        df_clean['تاریخ'] = df_clean['تاریخ'].ffill()
        
        df_clean = df_clean[['کد پرسنلی', 'نام و نام خانوادگی', 'تاریخ', 'ورود', 'گیت ورود', 'خروج', 'گیت خروج']]
        
        # ۳. پردازش ترددها و جداسازی ردیف‌های ناقص
        def count_not_na(x):
            return x.notna().sum()
            
        # شمارش تعداد ورود و خروج هر شخص در هر تاریخ
        counts = df_clean.groupby(['کد پرسنلی', 'تاریخ']).agg(
            Entries=('ورود', count_not_na),
            Exits=('خروج', count_not_na)
        ).reset_index()
        
        # شناسایی ترددهای دارای بیشتر از ۱ ورود یا بیشتر از ۱ خروج
        defective_groups = counts[(counts['Entries'] > 1) | (counts['Exits'] > 1)]
        healthy_groups = counts[~((counts['Entries'] > 1) | (counts['Exits'] > 1))]
        
        defective_df = pd.merge(df_clean, defective_groups[['کد پرسنلی', 'تاریخ']], on=['کد پرسنلی', 'تاریخ'])
        healthy_df = pd.merge(df_clean, healthy_groups[['کد پرسنلی', 'تاریخ']], on=['کد پرسنلی', 'تاریخ'])
        
        st.success(f"پردازش انجام شد! {len(defective_df)} رکورد استخراج شده (ناقص) و {len(healthy_df)} رکورد سالم یافت شد.")
        
        # توابع خروجی اکسل با فرمت‌دهی
        def to_excel_highlighted(df):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Defective')
                worksheet = writer.sheets['Defective']
                # اعمال هایلایت زرد روی تمام سلول‌ها
                yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.fill = yellow_fill
            return output.getvalue()

        def to_excel_normal(df):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Normal')
            return output.getvalue()
            
        col1, col2 = st.columns(2)
        
        with col1:
            st.warning("ترددهای ناقص (آماده استخراج)")
            st.dataframe(defective_df.head(10))
            st.download_button(
                label="دانلود فایل ترددهای استخراج‌شده",
                data=to_excel_highlighted(defective_df),
                file_name="Defective_Traffic.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        with col2:
            st.success("ترددهای سالم")
            st.dataframe(healthy_df.head(10))
            st.download_button(
                label="دانلود فایل ردیف‌های فاقد نقص",
                data=to_excel_normal(healthy_df),
                file_name="Healthy_Traffic.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"خطایی در پردازش فایل رخ داد: {e}")
