import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime, timedelta

st.set_page_config(page_title="محاسبه کارکرد پرسنل", layout="wide")

st.title("📊 سیستم استخراج و محاسبه کارکرد پرسنل")
st.write("لطفاً فایل اکسل تردد را آپلود کنید تا محاسبات کارکرد به صورت خودکار انجام شود.")

# آپلود فایل اکسل
uploaded_file = st.file_uploader("انتخاب فایل اکسل", type=["xlsx", "xls"])

def parse_time(time_str):
    """تبدیل رشته ساعت به شیء datetime"""
    if not time_str or pd.isna(time_str) or str(time_str).strip() in ['-', '']:
        return None
    
    # پاک‌سازی کاراکترهای مخفی و فاصله‌ها
    clean_str = str(time_str).replace('\u200f', '').replace('\u200e', '').strip()
    
    # تبدیل اعداد فارسی به انگلیسی
    persian_digits = '۰۱۲۳۴۵۶۷۸۹'
    english_digits = '0123456789'
    translation_table = str.maketrans(persian_digits, english_digits)
    clean_str = clean_str.translate(translation_table)
    
    try:
        return datetime.strptime(clean_str, "%H:%M")
    except ValueError:
        try:
            return datetime.strptime(clean_str, "%H:%M:%S")
        except ValueError:
            return None

def calculate_work_duration(entry_str, exit_str):
    """محاسبه اختلاف ساعت بر اساس شرط ساعت شروع رسمی (07:15)"""
    entry_dt = parse_time(entry_str)
    exit_dt = parse_time(exit_str)
    
    # اگر ورود یا خروج ثبت نشده باشد
    if not entry_dt or not exit_dt:
        return "-"
    
    # ساعت مبنای رسمی شرکت: ۰۷:۱۵
    official_start = entry_dt.replace(hour=7, minute=15, second=0)
    
    # اعمال شرط: اگر زودتر از 07:15 آمده باشد، 07:15 لحاظ می‌شود
    effective_entry = max(entry_dt, official_start)
    
    # محاسبه اختلاف زمانی (اگر خروج قبل از ورود باشد فرض بر عبور از نیمه‌شب است)
    if exit_dt < effective_entry:
        exit_dt += timedelta(days=1)
        
    duration = exit_dt - effective_entry
    
    total_seconds = int(duration.total_seconds())
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    
    return f"{hours:02d}:{minutes:02d}"

if uploaded_file is not None:
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        sheet = wb.active
        
        data = []
        # پیمایش سطرها از ردیف ۱۵ (طبق ساختار فایل)
        for r in range(15, sheet.max_row + 1):
            p_code = sheet.cell(row=r, column=20).value
            name = sheet.cell(row=r, column=19).value
            entry = sheet.cell(row=r, column=10).value
            exit_time = sheet.cell(row=r, column=5).value
            
            if p_code or name:
                entry_val = str(entry).strip() if entry is not None else '-'
                exit_val = str(exit_time).strip() if exit_time is not None else '-'
                
                # محاسبه مدت حضور
                duration = calculate_work_duration(entry_val, exit_val)
                
                data.append({
                    'کد پرسنلی': str(p_code).strip() if p_code is not None else '-',
                    'نام و نام خانوادگی': str(name).strip() if name is not None else '-',
                    'ورود': entry_val,
                    'خروج': exit_val,
                    'مدت حضور (کارکرد)': duration
                })
        
        df = pd.DataFrame(data)
        
        st.success(f"تعداد {len(df)} رکورد با موفقیت پردازش شد.")
        
        # نمایش جدول داده‌ها
        st.dataframe(df, use_container_width=True)
        
        # امکان دانلود خروجی به صورت اکسل جدید
        @st.cache_data
        def convert_df_to_excel(dataframe):
            import io
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                dataframe.to_excel(writer, index=False, sheet_name='گزارش کارکرد')
            return output.getvalue()
        
        excel_data = convert_df_to_excel(df)
        st.download_button(
            label="📥 دانلود فایل محاسبات (Excel)",
            data=excel_data,
            file_name="گزارش_محاسبه_کارکرد.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        st.error(f"خطا در پردازش فایل: {e}")
