import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta
import io
import os

st.set_page_config(page_title="سوپر اپ محاسبه و تجمیع اضافه کاری", layout="wide")

st.title("🚀 سوپر اپلیکیشن محاسبه و تجمیع اضافه کاری پرسنل")
st.write("فایل‌های اکسل تردد روزانه را آپلود کنید تا محاسبات فردی و گزارش تجمیعی کل با رعایت دقیق قوانین محاسبه گردد.")

# ----------------------------------------------------
# توابع محاسباتی و کمکی
# ----------------------------------------------------

def parse_time(time_str):
    """تبدیل رشته ساعت به شیء datetime با پشتیبانی از اعداد فارسی و انگلیسی"""
    if not time_str or pd.isna(time_str) or str(time_str).strip() in ['-', '']:
        return None
    
    clean_str = str(time_str).replace('\u200f', '').replace('\u200e', '').strip()
    
    persian_digits = '۰۱۲۳۴۵۶۷۸۹'
    english_digits = '0123456789'
    translation_table = str.maketrans(persian_digits, english_digits)
    clean_str = clean_str.translate(translation_table)
    
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(clean_str, fmt)
        except ValueError:
            pass
    return None

def calculate_duration_minutes(entry_str, exit_str):
    """
    محاسبه مدت کارکرد به دقیقه با اعمال شرط شروع رسمی (07:15)
    """
    entry_dt = parse_time(entry_str)
    exit_dt = parse_time(exit_str)
    
    if not entry_dt or not exit_dt:
        return 0
    
    # ساعت مبنای رسمی: 07:15
    official_start = entry_dt.replace(hour=7, minute=15, second=0)
    
    # شرط: ورود زودتر از 07:15 همان 07:15 محاسبه می‌شود
    effective_entry = max(entry_dt, official_start)
    
    if exit_dt < effective_entry:
        exit_dt += timedelta(days=1)
        
    duration = exit_dt - effective_entry
    return int(duration.total_seconds() // 60)

def minutes_to_hhmm(total_minutes):
    """تبدیل دقیقه به فرمت HH:MM"""
    if total_minutes <= 0:
        return "00:00"
    hours = total_minutes // 60
    minutes = total_minutes % 60
    return f"{hours:02d}:{minutes:02d}"

def process_single_excel(file_bytes, file_name):
    """پردازش یک فایل اکسل و استخراج داده‌ها و محاسبات"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = wb.active
    
    records = []
    for r in range(15, sheet.max_row + 1):
        p_code = sheet.cell(row=r, column=20).value
        name = sheet.cell(row=r, column=19).value
        entry = sheet.cell(row=r, column=10).value
        exit_time = sheet.cell(row=r, column=5).value
        
        if p_code or name:
            p_code_str = str(p_code).strip() if p_code is not None else '-'
            name_str = str(name).strip() if name is not None else '-'
            entry_str = str(entry).strip() if entry is not None else '-'
            exit_str = str(exit_time).strip() if exit_time is not None else '-'
            
            dur_min = calculate_duration_minutes(entry_str, exit_str)
            dur_hhmm = minutes_to_hhmm(dur_min)
            
            records.append({
                'کد پرسنلی': p_code_str,
                'نام و نام خانوادگی': name_str,
                'ورود': entry_str,
                'خروج': exit_str,
                'کارکرد (دقیقه)': dur_min,
                'مدت حضور (کارکرد)': dur_hhmm
            })
            
    df = pd.DataFrame(records)
    return df

def generate_styled_master_excel(master_df, file_columns):
    """تولید فایل اکسل نهایی تجمیعی با هایلایت و استایل حرفه‌ای"""
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "گزارش تجمیعی کارکرد"
    ws.views.sheetView[0].rightToLeft = True  # راست به چپ برای فارسی
    
    headers = list(master_df.columns)
    ws.append(headers)
    
    # استایل هدر
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # استایل‌های بدنه
    highlight_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # سبز ملایم
    zero_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    regular_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", size=10, bold=True)
    
    for row_idx, row_data in enumerate(master_df.values, 2):
        has_work = False
        total_work_val = row_data[-1]  # ستون آخر "مجموع کارکرد کل"
        
        if total_work_val != "00:00" and total_work_val != "-":
            has_work = True
            
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = bold_font if col_idx in [1, 2, len(row_data)] else regular_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
            # هایلایت سطرهایی که کارکرد دارند
            if has_work:
                cell.fill = highlight_fill
                
    # تنظیم عرض ستون‌ها
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    wb.save(output)
    return output.getvalue()

# ----------------------------------------------------
# رابط کاربری Streamlit
# ----------------------------------------------------

uploaded_files = st.file_uploader(
    "فایل‌های اکسل تردد را انتخاب یا در این محل رها کنید (امکان انتخاب چندتایی):",
    type=["xlsx", "xls"],
    accept_multiple_files=True
)

if uploaded_files:
    st.divider()
    st.subheader("📋 فایل‌های پردازش‌شده به تفکیک:")
    
    processed_dfs = {}
    personnel_info = {} # برای ذخیره نام و نام خانوادگی کد پرسنلی‌ها
    
    cols = st.columns(min(len(uploaded_files), 3))
    
    for idx, file in enumerate(uploaded_files):
        file_bytes = file.read()
        file_clean_name = os.path.splitext(file.name)[0]
        
        # پردازش فایل
        df_single = process_single_excel(file_bytes, file.name)
        processed_dfs[file_clean_name] = df_single
        
        # ذخیره نگاشت کد پرسنلی به نام
        for _, r in df_single.iterrows():
            p_code = r['کد پرسنلی']
            if p_code != '-' and p_code not in personnel_info:
                personnel_info[p_code] = r['نام و نام خانوادگی']
                
        # نمایش کارت خلاصه در UI
        with cols[idx % len(cols)]:
            st.info(f"📄 **{file.name}**")
            st.write(f"تعداد پرسنل: {len(df_single)}")
            
            # خروجی اکسل تک‌تک فایل‌ها
            output_single = io.BytesIO()
            with pd.ExcelWriter(output_single, engine='openpyxl') as writer:
                df_single.drop(columns=['کارکرد (دقیقه)']).to_excel(writer, index=False)
                
            st.download_button(
                label=f"📥 دانلود پردازش {file_clean_name}",
                data=output_single.getvalue(),
                file_name=f"محاسبه_{file.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_{idx}"
            )

    # ----------------------------------------------------
    # ساخت جدول تجمیعی نهایی (Merge)
    # ----------------------------------------------------
    st.divider()
    st.subheader("📊 گزارش تجمیعی کل (Merge کامل بر اساس کد پرسنلی)")
    
    all_p_codes = sorted(list(personnel_info.keys()))
    
    master_rows = []
    file_names = list(processed_dfs.keys())
    
    for p_code in all_p_codes:
        row_dict = {
            'کد پرسنلی': p_code,
            'نام و نام خانوادگی': personnel_info.get(p_code, '-')
        }
        
        total_minutes = 0
        
        for f_name in file_names:
            df_file = processed_dfs[f_name]
            match = df_file[df_file['کد پرسنلی'] == p_code]
            
            if not match.empty:
                dur_min = match.iloc[0]['کارکرد (دقیقه)']
                dur_str = match.iloc[0]['مدت حضور (کارکرد)']
                total_minutes += dur_min
                row_dict[f"کارکرد - {f_name}"] = dur_str
            else:
                row_dict[f"کارکرد - {f_name}"] = "-"
                
        row_dict['مجموع کارکرد کل'] = minutes_to_hhmm(total_minutes)
        master_rows.append(row_dict)
        
    master_df = pd.DataFrame(master_rows)
    
    # نمایش جدول تجمیعی
    st.dataframe(master_df, use_container_width=True)
    
    # تولید و دانلود فایل تجمیعی شکیل با هایلایت
    master_excel_bytes = generate_styled_master_excel(master_df, file_names)
    
    st.download_button(
        label="🔥 📥 دانلود اکسل تجمیعی کل (با هایلایت و استایل مدیریتی)",
        data=master_excel_bytes,
        file_name="گزارش_تجمیعی_اضافه_کاری_کل.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
